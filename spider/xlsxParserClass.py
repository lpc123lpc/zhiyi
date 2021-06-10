import openpyxl
import pandas as pd
class xlsxParser:
    def __init__(self,path):
        self.path=path
        self.workbook = openpyxl.load_workbook(path)
        self.sheet=self.workbook[self.workbook.sheetnames[0]]
        self.max_row = self.sheet.max_row
        self.max_col = self.sheet.max_column
        self.mergedCellsDict= {}
        self.updateMergedCellsDict()

    def changeSheet(self,sheetName):
        self.sheet=self.workbook[sheetName]
        self.max_row=self.sheet.max_row
        self.max_col=self.sheet.max_column
        self.mergedCellsDict.clear()
        self.updateMergedCellsDict()

    def setMaxRowCol(self,max_row,max_col):
        self.max_row=max_row
        self.max_col=max_col

    def displayConfig(self):
        print("myXslx: "+self.path)
        print("mySheet: "+self.sheet.title)
        print("max_row: "+str(self.max_row))
        print("max_col: "+str(self.max_col))
        print(self.mergedCellsDict)

    def updateMergedCellsDict(self):
        for item in self.sheet.merged_cells:
            c1,r1,c2,r2=item.bounds
            mergedValue=self.sheet.cell(r1,c1).value
            if mergedValue==None:
                mergedValue="暂未公布"
            print("mergedValue: ", self.sheet.cell(r1,c1).value)
            for mergeRow in range(r1,r2+1):
                for mergeCol in range(c1,c2+1):
                    if(mergeRow,mergeCol)!=(r1,c1):
                        self.mergedCellsDict[(mergeRow,mergeCol)]=mergedValue

    def getValue(self,row,col):
        if self.sheet.cell(row,col).value==None:
            if self.mergedCellsDict.get((row,col))==None:
                print("Exists Blank: ("+str(row)+","+str(col)+") "+self.sheet.title)
                return None
            else:
                return self.mergedCellsDict.get((row,col))
        else:
            return self.sheet.cell(row,col).value
